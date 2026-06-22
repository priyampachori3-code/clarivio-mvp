import io
import json
import os
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from fastapi import Depends, FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from passlib.context import CryptContext
from pydantic import BaseModel
from rapidfuzz import fuzz
from starlette.staticfiles import StaticFiles

SECRET_KEY = os.getenv('SECRET_KEY', 'clarivio-secret-change-in-production')
ALGORITHM = 'HS256'
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 8
USERS_DB = {}
pwd_context = CryptContext(schemes=['bcrypt'], deprecated='auto')
oauth2_scheme = OAuth2PasswordBearer(tokenUrl='/auth/login')

app = FastAPI(title='Clarivio API', version='1.0.0')
origins = [origin.strip() for origin in os.getenv('FRONTEND_ORIGINS', 'http://localhost:5173').split(',')]
allow_origins = ['*'] if '*' in origins else origins
app.add_middleware(
    CORSMiddleware,
    allow_origins=allow_origins,
    allow_credentials=False if allow_origins == ['*'] else True,
    allow_methods=['*'],
    allow_headers=['*'],
)

class RegisterRequest(BaseModel):
    email: str
    password: str
    firm_name: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str
    firm_name: str

def hash_password(password: str) -> str:
    return pwd_context.hash(password[:72])

def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain[:72], hashed)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    to_encode.update({'exp': datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def decode_token(token: str) -> str | None:
    try:
        return jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM]).get('sub')
    except JWTError:
        return None

def get_current_user(token: str = Depends(oauth2_scheme)):
    email = decode_token(token)
    if not email or email not in USERS_DB:
        raise HTTPException(status_code=401, detail='Invalid or expired token')
    return USERS_DB[email]

@app.post('/auth/register', response_model=TokenResponse)
def register(data: RegisterRequest):
    if data.email in USERS_DB:
        raise HTTPException(status_code=400, detail='Email already registered')
    USERS_DB[data.email] = {'email': data.email, 'hashed_password': hash_password(data.password), 'firm_name': data.firm_name}
    token = create_access_token({'sub': data.email})
    return {'access_token': token, 'token_type': 'bearer', 'firm_name': data.firm_name}

@app.post('/auth/login', response_model=TokenResponse)
def login(form_data: OAuth2PasswordRequestForm = Depends()):
    user = USERS_DB.get(form_data.username)
    if not user or not verify_password(form_data.password, user['hashed_password']):
        raise HTTPException(status_code=401, detail='Invalid credentials')
    token = create_access_token({'sub': form_data.username})
    return {'access_token': token, 'token_type': 'bearer', 'firm_name': user['firm_name']}

@app.get('/auth/me')
def me(current_user: dict = Depends(get_current_user)):
    return {'email': current_user['email'], 'firm_name': current_user['firm_name']}

def norm(value) -> str:
    return str(value or '').strip().upper().replace(' ', '').replace('-', '')

def amount(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0

def parse_gstr2b(file_bytes: bytes) -> pd.DataFrame:
    data = json.loads(file_bytes)
    b2b = data.get('data', {}).get('docdata', {}).get('b2b', data.get('b2b', []))
    records = []
    for supplier in b2b:
        gstin = supplier.get('ctin', '')
        for inv in supplier.get('inv', []):
            item = (inv.get('itms') or [{}])[0].get('itm_det', {})
            records.append({
    'gstin': norm(gstin),
    'invoice_no': norm(inv.get('inum', '')),
    'invoice_date': inv.get('idt', ''),
    'taxable_value': amount(inv.get('taxval', 0)),
    # fix - read directly from invoice
    'igst': amount(inv.get('igst') or item.get('iamt', 0)),
    'cgst': amount(inv.get('cgst') or item.get('camt', 0)),
    'sgst': amount(inv.get('sgst') or item.get('samt', 0)),
})
    return pd.DataFrame(records)

def parse_tally_excel(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes), engine='openpyxl')
    df.columns = [str(c).strip().lower() for c in df.columns]
    candidates = {
        'gstin': ['gstin', 'party gstin', 'supplier gstin'],
        'invoice_no': ['invoice no', 'invoice number', 'inv no', 'voucher no'],
        'invoice_date': ['invoice date', 'inv date', 'date'],
        'taxable_value': ['taxable value', 'taxable amount', 'basic amount'],
        'igst': ['igst', 'igst amount'],
        'cgst': ['cgst', 'cgst amount'],
        'sgst': ['sgst', 'sgst amount'],
    }
    rename = {}
    for target, names in candidates.items():
        for name in names:
            if name in df.columns:
                rename[name] = target
                break
    df = df.rename(columns=rename)
    for col in candidates:
        if col not in df.columns:
            df[col] = '' if col in {'gstin', 'invoice_no', 'invoice_date'} else 0
    df['gstin'] = df['gstin'].apply(norm)
    df['invoice_no'] = df['invoice_no'].apply(norm)
    for col in ['taxable_value', 'igst', 'cgst', 'sgst']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    return df[['gstin', 'invoice_no', 'invoice_date', 'taxable_value', 'igst', 'cgst', 'sgst']]

def merged_row(portal, tally):
    return {
        'gstin': portal['gstin'],
        'invoice_no_portal': portal['invoice_no'],
        'invoice_no_tally': tally['invoice_no'],
        'invoice_date_portal': portal['invoice_date'],
        'invoice_date_tally': tally['invoice_date'],
        'taxable_portal': portal['taxable_value'],
        'taxable_tally': tally['taxable_value'],
        'igst_portal': portal['igst'],
        'igst_tally': tally['igst'],
        'cgst_portal': portal['cgst'],
        'cgst_tally': tally['cgst'],
        'sgst_portal': portal['sgst'],
        'sgst_tally': tally['sgst'],
        'diff_taxable': round(portal['taxable_value'] - tally['taxable_value'], 2),
    }

def single_row(row, source):
    return {
        'gstin': row['gstin'],
        'invoice_no_portal': row['invoice_no'] if source == 'portal' else '',
        'invoice_no_tally': row['invoice_no'] if source == 'tally' else '',
        'invoice_date_portal': row['invoice_date'] if source == 'portal' else '',
        'invoice_date_tally': row['invoice_date'] if source == 'tally' else '',
        'taxable_portal': row['taxable_value'] if source == 'portal' else 0,
        'taxable_tally': row['taxable_value'] if source == 'tally' else 0,
        'igst_portal': row['igst'] if source == 'portal' else 0,
        'igst_tally': row['igst'] if source == 'tally' else 0,
        'cgst_portal': row['cgst'] if source == 'portal' else 0,
        'cgst_tally': row['cgst'] if source == 'tally' else 0,
        'sgst_portal': row['sgst'] if source == 'portal' else 0,
        'sgst_tally': row['sgst'] if source == 'tally' else 0,
        'diff_taxable': 0,
    }

def reconcile(portal_df: pd.DataFrame, tally_df: pd.DataFrame) -> dict:
    results = []
    portal = {(r['gstin'], r['invoice_no']): r for _, r in portal_df.iterrows()}
    tally = {(r['gstin'], r['invoice_no']): r for _, r in tally_df.iterrows()}
    matched_portal, matched_tally = set(), set()
    for key, p_row in portal.items():
        if key in tally:
            t_row = tally[key]
            status = 'MATCHED' if abs(p_row['taxable_value'] - t_row['taxable_value']) <= 1 else 'AMOUNT_MISMATCH'
            results.append({**merged_row(p_row, t_row), 'status': status, 'match_score': 100})
            matched_portal.add(key)
            matched_tally.add(key)
    unmatched_tally = {k: v for k, v in tally.items() if k not in matched_tally}
    for p_key, p_row in [(k, v) for k, v in portal.items() if k not in matched_portal]:
        best_key, best_score = None, 0
        for t_key, t_row in unmatched_tally.items():
            if p_row['gstin'] == t_row['gstin']:
                score = fuzz.ratio(p_row['invoice_no'], t_row['invoice_no'])
                if score > best_score:
                    best_key, best_score = t_key, score
        if best_key and best_score >= 85:
            t_row = unmatched_tally.pop(best_key)
            results.append({**merged_row(p_row, t_row), 'status': 'FUZZY_MATCH', 'match_score': best_score})
            matched_portal.add(p_key)
            matched_tally.add(best_key)
    for key, row in portal.items():
        if key not in matched_portal:
            results.append({**single_row(row, 'portal'), 'status': 'PORTAL_ONLY', 'match_score': 0})
    for key, row in tally.items():
        if key not in matched_tally:
            results.append({**single_row(row, 'tally'), 'status': 'TALLY_ONLY', 'match_score': 0})
    total = len(results)
    matched = sum(1 for r in results if r['status'] == 'MATCHED')
    mismatch = sum(1 for r in results if r['status'] in ('AMOUNT_MISMATCH', 'FUZZY_MATCH'))
    portal_only = sum(1 for r in results if r['status'] == 'PORTAL_ONLY')
    tally_only = sum(1 for r in results if r['status'] == 'TALLY_ONLY')
    return {'summary': {'total': total, 'matched': matched, 'mismatch': mismatch, 'portal_only': portal_only, 'tally_only': tally_only, 'match_rate': round(matched / total * 100, 1) if total else 0}, 'rows': results}

def generate_excel_report(result: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Reconciliation'
    headers = ['Status', 'GSTIN', 'Invoice Portal', 'Invoice Tally', 'Taxable Portal', 'Taxable Tally', 'Diff']
    ws.append(headers)
    fill_map = {'MATCHED': 'C6EFCE', 'AMOUNT_MISMATCH': 'FFEB9C', 'FUZZY_MATCH': 'FFEB9C', 'PORTAL_ONLY': 'FFC7CE', 'TALLY_ONLY': 'BDD7EE'}
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E79')
    for row in result['rows']:
        ws.append([row['status'], row['gstin'], row['invoice_no_portal'], row['invoice_no_tally'], row['taxable_portal'], row['taxable_tally'], row['diff_taxable']])
        fill = PatternFill('solid', fgColor=fill_map.get(row['status'], 'FFFFFF'))
        for cell in ws[ws.max_row]:
            cell.fill = fill
    for index, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(index)].width = max(14, len(header) + 2)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

@app.post('/reconcile/run')
async def run_reconciliation(gstr2b_file: UploadFile = File(...), tally_file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    try:
        portal_df = parse_gstr2b(await gstr2b_file.read())
        tally_df = parse_tally_excel(await tally_file.read())
        return reconcile(portal_df, tally_df)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))

@app.post('/reconcile/download')
async def download_report(gstr2b_file: UploadFile = File(...), tally_file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    try:
        result = reconcile(parse_gstr2b(await gstr2b_file.read()), parse_tally_excel(await tally_file.read()))
        return Response(content=generate_excel_report(result), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment; filename=clarivio_reconciliation.xlsx'})
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))

@app.get('/api/health')
def health():
    return {'status': 'Clarivio API running'}

frontend_dist = Path(__file__).resolve().parent.parent / 'frontend' / 'dist'
if frontend_dist.exists():
    app.mount('/assets', StaticFiles(directory=frontend_dist / 'assets'), name='assets')

    @app.get('/{full_path:path}', include_in_schema=False)
    def serve_frontend(full_path: str):
        requested = frontend_dist / full_path
        if full_path and requested.is_file():
            return FileResponse(requested)
        return FileResponse(frontend_dist / 'index.html')

@app.get('/')
def root():
    return {'status': 'Clarivio API running'}
