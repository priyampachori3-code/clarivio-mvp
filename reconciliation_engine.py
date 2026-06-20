"""
Clarivio GST Reconciliation Engine
Matches GSTR-2B (portal JSON) vs Tally purchase register (Excel)
"""
import json
import io
import pandas as pd
from rapidfuzz import fuzz
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ──────────────────────────────────────────────────────────
GREEN  = "C6EFCE"   # matched
YELLOW = "FFEB9C"   # fuzzy / amount mismatch
RED    = "FFC7CE"   # in portal, not in Tally
BLUE   = "BDD7EE"   # in Tally, not in portal
HEADER = "1F4E79"   # deep navy header

def _norm(s: str) -> str:
    """Normalise GSTIN / invoice numbers for comparison."""
    return str(s).strip().upper().replace(" ", "").replace("-", "")

def parse_gstr2b(file_bytes: bytes) -> pd.DataFrame:
    """Parse GSTR-2B JSON downloaded from GST portal."""
    data = json.loads(file_bytes)
    records = []
    # Standard GSTR-2B structure: data > docdata > b2b
    try:
        b2b = data["data"]["docdata"]["b2b"]
    except KeyError:
        # Fallback: try root-level b2b key
        b2b = data.get("b2b", [])

    for supplier in b2b:
        gstin = supplier.get("ctin", "")
        for inv in supplier.get("inv", []):
            records.append({
                "source":        "portal",
                "gstin":         _norm(gstin),
                "invoice_no":    _norm(inv.get("inum", "")),
                "invoice_date":  inv.get("idt", ""),
                "taxable_value": float(inv.get("val", 0)),
                "igst":          float(inv.get("itms", [{}])[0].get("itm_det", {}).get("iamt", 0)),
                "cgst":          float(inv.get("itms", [{}])[0].get("itm_det", {}).get("camt", 0)),
                "sgst":          float(inv.get("itms", [{}])[0].get("itm_det", {}).get("samt", 0)),
            })
    return pd.DataFrame(records)

def parse_tally_excel(file_bytes: bytes) -> pd.DataFrame:
    """
    Parse Tally purchase register Excel.
    Expects columns (case-insensitive):
      GSTIN / Party GSTIN, Invoice No, Invoice Date, Taxable Value, IGST, CGST, SGST
    """
    df = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")
    df.columns = [c.strip().lower() for c in df.columns]

    col_map = {
        "gstin":         ["gstin", "party gstin", "supplier gstin"],
        "invoice_no":    ["invoice no", "invoice number", "inv no", "voucher no"],
        "invoice_date":  ["invoice date", "inv date", "date"],
        "taxable_value": ["taxable value", "taxable amount", "basic amount"],
        "igst":          ["igst", "igst amount"],
        "cgst":          ["cgst", "cgst amount"],
        "sgst":          ["sgst", "sgst amount"],
    }

    rename = {}
    for target, candidates in col_map.items():
        for c in candidates:
            if c in df.columns:
                rename[c] = target
                break

    df = df.rename(columns=rename)

    for col in ["gstin", "invoice_no", "invoice_date", "taxable_value", "igst", "cgst", "sgst"]:
        if col not in df.columns:
            df[col] = "" if col in ["gstin", "invoice_no", "invoice_date"] else 0.0

    df["gstin"]      = df["gstin"].apply(_norm)
    df["invoice_no"] = df["invoice_no"].apply(_norm)
    df["taxable_value"] = pd.to_numeric(df["taxable_value"], errors="coerce").fillna(0)
    df["igst"]  = pd.to_numeric(df["igst"],  errors="coerce").fillna(0)
    df["cgst"]  = pd.to_numeric(df["cgst"],  errors="coerce").fillna(0)
    df["sgst"]  = pd.to_numeric(df["sgst"],  errors="coerce").fillna(0)
    df["source"] = "tally"

    return df[["source", "gstin", "invoice_no", "invoice_date",
               "taxable_value", "igst", "cgst", "sgst"]]

def reconcile(portal_df: pd.DataFrame, tally_df: pd.DataFrame) -> dict:
    """
    Match invoices between portal and Tally.
    Returns summary dict + detailed result rows.
    """
    AMOUNT_TOLERANCE = 1.0   # ₹1 rounding tolerance
    FUZZY_THRESHOLD  = 85    # RapidFuzz score threshold

    results = []

    portal_dict = {
        (r["gstin"], r["invoice_no"]): r
        for _, r in portal_df.iterrows()
    }
    tally_dict = {
        (r["gstin"], r["invoice_no"]): r
        for _, r in tally_df.iterrows()
    }

    matched_portal_keys = set()
    matched_tally_keys  = set()

    # ── Exact match pass ────────────────────────────────────────────────────
    for key, p_row in portal_dict.items():
        if key in tally_dict:
            t_row = tally_dict[key]
            amount_ok = abs(p_row["taxable_value"] - t_row["taxable_value"]) <= AMOUNT_TOLERANCE
            status = "MATCHED" if amount_ok else "AMOUNT_MISMATCH"
            results.append({
                **_merge_row(p_row, t_row),
                "status": status,
                "match_score": 100,
            })
            matched_portal_keys.add(key)
            matched_tally_keys.add(key)

    # ── Fuzzy match pass (for unmatched) ────────────────────────────────────
    unmatched_portal = {k: v for k, v in portal_dict.items() if k not in matched_portal_keys}
    unmatched_tally  = {k: v for k, v in tally_dict.items()  if k not in matched_tally_keys}

    for p_key, p_row in list(unmatched_portal.items()):
        best_score = 0
        best_t_key = None
        for t_key, t_row in unmatched_tally.items():
            if p_row["gstin"] != t_row["gstin"]:
                continue
            score = fuzz.ratio(p_row["invoice_no"], t_row["invoice_no"])
            if score > best_score:
                best_score = score
                best_t_key = t_key

        if best_t_key and best_score >= FUZZY_THRESHOLD:
            t_row = unmatched_tally[best_t_key]
            results.append({
                **_merge_row(p_row, t_row),
                "status": "FUZZY_MATCH",
                "match_score": best_score,
            })
            matched_portal_keys.add(p_key)
            matched_tally_keys.add(best_t_key)
            del unmatched_tally[best_t_key]

    # ── In portal, not in Tally ──────────────────────────────────────────────
    for p_key, p_row in portal_dict.items():
        if p_key not in matched_portal_keys:
            results.append({
                **_single_row(p_row, "portal"),
                "status": "PORTAL_ONLY",
                "match_score": 0,
            })

    # ── In Tally, not in portal ──────────────────────────────────────────────
    for t_key, t_row in tally_dict.items():
        if t_key not in matched_tally_keys:
            results.append({
                **_single_row(t_row, "tally"),
                "status": "TALLY_ONLY",
                "match_score": 0,
            })

    total     = len(results)
    matched   = sum(1 for r in results if r["status"] == "MATCHED")
    mismatch  = sum(1 for r in results if r["status"] in ("AMOUNT_MISMATCH", "FUZZY_MATCH"))
    portal_only = sum(1 for r in results if r["status"] == "PORTAL_ONLY")
    tally_only  = sum(1 for r in results if r["status"] == "TALLY_ONLY")

    return {
        "summary": {
            "total":        total,
            "matched":      matched,
            "mismatch":     mismatch,
            "portal_only":  portal_only,
            "tally_only":   tally_only,
            "match_rate":   round(matched / total * 100, 1) if total else 0,
        },
        "rows": results,
    }

def _merge_row(p, t) -> dict:
    return {
        "gstin":               p["gstin"],
        "invoice_no_portal":   p["invoice_no"],
        "invoice_no_tally":    t["invoice_no"],
        "invoice_date_portal": p["invoice_date"],
        "invoice_date_tally":  t["invoice_date"],
        "taxable_portal":      p["taxable_value"],
        "taxable_tally":       t["taxable_value"],
        "igst_portal":         p["igst"],
        "igst_tally":          t["igst"],
        "cgst_portal":         p["cgst"],
        "cgst_tally":          t["cgst"],
        "sgst_portal":         p["sgst"],
        "sgst_tally":          t["sgst"],
        "diff_taxable":        round(p["taxable_value"] - t["taxable_value"], 2),
    }

def _single_row(r, source: str) -> dict:
    empty = "" if source == "portal" else r["invoice_no"]
    portal_inv = r["invoice_no"] if source == "portal" else ""
    tally_inv  = r["invoice_no"] if source == "tally"  else ""
    return {
        "gstin":               r["gstin"],
        "invoice_no_portal":   portal_inv,
        "invoice_no_tally":    tally_inv,
        "invoice_date_portal": r["invoice_date"] if source == "portal" else "",
        "invoice_date_tally":  r["invoice_date"] if source == "tally"  else "",
        "taxable_portal":      r["taxable_value"] if source == "portal" else 0,
        "taxable_tally":       r["taxable_value"] if source == "tally"  else 0,
        "igst_portal":         r["igst"] if source == "portal" else 0,
        "igst_tally":          r["igst"] if source == "tally"  else 0,
        "cgst_portal":         r["cgst"] if source == "portal" else 0,
        "cgst_tally":          r["cgst"] if source == "tally"  else 0,
        "sgst_portal":         r["sgst"] if source == "portal" else 0,
        "sgst_tally":          r["sgst"] if source == "tally"  else 0,
        "diff_taxable":        0,
    }

def generate_excel_report(result: dict) -> bytes:
    """Generate a colour-coded Excel reconciliation report."""
    wb = Workbook()

    # ── Summary Sheet ────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_fill  = PatternFill("solid", fgColor=HEADER)
    header_font  = Font(color="FFFFFF", bold=True, size=12)
    bold_font    = Font(bold=True)
    center       = Alignment(horizontal="center", vertical="center")

    ws_summary.merge_cells("A1:B1")
    ws_summary["A1"] = "Clarivio — GST Reconciliation Summary"
    ws_summary["A1"].font  = Font(bold=True, size=14, color=HEADER)
    ws_summary["A1"].alignment = center

    summary_data = [
        ("Total Invoices",    result["summary"]["total"]),
        ("Matched",           result["summary"]["matched"]),
        ("Mismatches",        result["summary"]["mismatch"]),
        ("Portal Only",       result["summary"]["portal_only"]),
        ("Tally Only",        result["summary"]["tally_only"]),
        ("Match Rate (%)",    result["summary"]["match_rate"]),
    ]

    status_colours = {
        "Matched":       GREEN,
        "Mismatches":    YELLOW,
        "Portal Only":   RED,
        "Tally Only":    BLUE,
    }

    for i, (label, value) in enumerate(summary_data, start=3):
        ws_summary[f"A{i}"] = label
        ws_summary[f"B{i}"] = value
        ws_summary[f"A{i}"].font = bold_font
        colour = status_colours.get(label)
        if colour:
            ws_summary[f"B{i}"].fill = PatternFill("solid", fgColor=colour)

    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 16

    # ── Detail Sheet ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Reconciliation Detail")

    columns = [
        "Status", "Match Score", "GSTIN",
        "Invoice No (Portal)", "Invoice No (Tally)",
        "Date (Portal)", "Date (Tally)",
        "Taxable (Portal)", "Taxable (Tally)", "Diff (Taxable)",
        "IGST (Portal)", "IGST (Tally)",
        "CGST (Portal)", "CGST (Tally)",
        "SGST (Portal)", "SGST (Tally)",
    ]

    # Header row
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill  = PatternFill("solid", fgColor=HEADER)
        cell.font  = header_font
        cell.alignment = center

    STATUS_COLOUR = {
        "MATCHED":        GREEN,
        "AMOUNT_MISMATCH": YELLOW,
        "FUZZY_MATCH":    YELLOW,
        "PORTAL_ONLY":    RED,
        "TALLY_ONLY":     BLUE,
    }

    for row_idx, row in enumerate(result["rows"], start=2):
        colour = STATUS_COLOUR.get(row["status"], "FFFFFF")
        fill   = PatternFill("solid", fgColor=colour)
        values = [
            row["status"], row["match_score"], row["gstin"],
            row["invoice_no_portal"], row["invoice_no_tally"],
            row["invoice_date_portal"], row["invoice_date_tally"],
            row["taxable_portal"], row["taxable_tally"], row["diff_taxable"],
            row["igst_portal"], row["igst_tally"],
            row["cgst_portal"], row["cgst_tally"],
            row["sgst_portal"], row["sgst_tally"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill

    # Auto-width
    for col_idx, col_name in enumerate(columns, start=1):
        max_len = max(len(col_name), 12)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
